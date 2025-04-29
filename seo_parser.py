import re
import warnings
import openpyxl
from models import SEORequirements, HeadingTargets
from utils.logger import get_logger
from utils.errors import ParseError
import math


# logger setup
logger = get_logger(__name__)

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl.styles.stylesheet")

def extract_value(sheet, code, default=0):
    """
    Extract a value from a sheet based on a code.
    
    Args:
        sheet: The worksheet to extract from
        code: The code to look for in column 2
        default: Default value if not found or invalid
        
    Returns:
        int: The extracted value or default
    """
    for row in range(1, sheet.max_row + 1):
        if sheet.cell(row=row, column=2).value == code:
            value = sheet.cell(row=row, column=5).value
            if value:
                if isinstance(value, (int, float)):
                    return int(value)
                elif isinstance(value, str) and value.strip().isdigit():
                    return int(value.strip())
            return default
    return default

def parse_cora_report(file_path):
    """
    Parses a CORA Excel report and extracts SEO requirements.
    Restored to match the original extraction logic from backups/main.py, including custom entities and all original fields.
    """
    try:
        # Load the Excel workbook
        wb = openpyxl.load_workbook(file_path, data_only=True)
        
        # Initialize default values
        primary_keyword = ""
        entities = []
        custom_entities = []
        variations = []
        lsi_keywords = {}
        heading_structure = {}
        requirements = {}
        basic_tunings = {}
        word_count = 1500  # Default
        
        # Debug info
        debug_info = {
            "sheets_found": wb.sheetnames,
            "lsi_start_row": None,
            "entities_start_row": None,
            "headings_section": None
        }
        
        # Extract requirements from CORA report
        requirements = {}
        
        # Parse "Roadmap" sheet
        if "Roadmap" in wb.sheetnames:
            roadmap_sheet = wb["Roadmap"]
            
            # Variations from A2
            raw_variations = roadmap_sheet["A2"].value
            variations = [v.strip(' "\'') for v in raw_variations.split(",") if v.strip()] if raw_variations else []
            
            # Extract requirements from "Phase 1: Title & Headings" using regex
            marker_start_pattern = re.compile(r"Phase\s+1\b")
            possible_end_patterns = [
                re.compile(r"Phase\s+2\b"),
                re.compile(r"Phase\s+3\b"),
                re.compile(r"Phase\s+4\b"),
                re.compile(r"Phase\s+6\b"),
                re.compile(r"Phase\s+7\b"),
                re.compile(r"Phase\s+8\b"),
                re.compile(r"Phase\s+9\b"),
                re.compile(r"Phase\s+10\b")
            ]

            # Find start row using regex
            start_row = None
            for row in range(1, roadmap_sheet.max_row + 1):
                cell_a = roadmap_sheet.cell(row=row, column=1).value
                if cell_a and marker_start_pattern.search(str(cell_a).strip()):
                    start_row = row + 1
                    break
                    
            if start_row:
                # Find end row based on regex patterns
                end_row = None
                for row in range(start_row, roadmap_sheet.max_row + 1):
                    cell_a = roadmap_sheet.cell(row=row, column=1).value
                    if cell_a:
                        cell_text = str(cell_a).strip()
                        if any(pattern.search(cell_text) for pattern in possible_end_patterns):
                            end_row = row
                            break
                
                if not end_row:
                    end_row = roadmap_sheet.max_row
                
                # Extract requirements
                for row in range(start_row, end_row):
                    req_desc = roadmap_sheet.cell(row=row, column=1).value
                    req_amount_text = roadmap_sheet.cell(row=row, column=2).value
                    if req_desc and req_amount_text:
                        try:
                            req_amount_text_str = str(req_amount_text).strip()
                            add_more_match = re.search(r'Add\s+(\d+)\s+more', req_amount_text_str, re.IGNORECASE)
                            colon_match = re.search(r'[:=]\s*(\d+)', req_amount_text_str)
                            end_match = re.search(r'(\d+)\s*$', req_amount_text_str)
                            any_match = re.search(r'(\d+)', req_amount_text_str)
                            if add_more_match:
                                amount = int(add_more_match.group(1))
                            elif colon_match:
                                amount = int(colon_match.group(1))
                            elif end_match:
                                amount = int(end_match.group(1))
                            elif any_match:
                                amount = int(any_match.group(1))
                            else:
                                logger.warning(f"No number found in: {req_amount_text_str}")
                                continue
                            requirements[req_desc] = amount
                        except (ValueError, TypeError) as e:
                            logger.warning(f"Could not parse requirement amount: {req_amount_text}. Error: {str(e)}")
                            continue
        
        # Process "Basic Tunings" sheet in a single pass
        if "Basic Tunings" in wb.sheetnames:
            sheet = wb["Basic Tunings"]
            codes = {
                "CP492": "Word Count",
                "CPXR004": "Number of H1 tags",
                "CPXR005": "Number of H2 tags",
                "CPXR006": "Number of H3 tags",
                "CPXR007": "Number of H4 tags",
                "CPXR008": "Number of H5 tags",
                "CPXR009": "Number of H6 tags",
                "CP426": "Number of Images",
                "CPXR003": "Number of heading tags",
                "CP480": "Title Length",
                "CP380": "Description Length"
            }
            primary_keyword = sheet["B1"].value.strip() if sheet["B1"].value else ""
            basic_tunings["primary_keyword"] = primary_keyword
            basic_tunings["Word Count"] = extract_value(sheet, "CP492", 1500)
            basic_tunings["Number of H1 tags"] = extract_value(sheet, "CPXR004")
            basic_tunings["Number of H2 tags"] = extract_value(sheet, "CPXR005")
            basic_tunings["Number of H3 tags"] = extract_value(sheet, "CPXR006")
            basic_tunings["Number of H4 tags"] = extract_value(sheet, "CPXR007")
            basic_tunings["Number of H5 tags"] = extract_value(sheet, "CPXR008")
            basic_tunings["Number of H6 tags"] = extract_value(sheet, "CPXR009")
            basic_tunings["Number of Images"] = extract_value(sheet, "CP426")
            basic_tunings["Number of heading tags"] = extract_value(sheet, "CPXR003")
            requirements["Title Length"] = extract_value(sheet, "CP480", 60)
            requirements["Description Length"] = extract_value(sheet, "CP380", 160)
        
        # Set defaults for any missing values
        basic_tunings.setdefault("Word Count", 1500)
        basic_tunings.setdefault("Number of H1 tags", 0)
        basic_tunings.setdefault("Number of H2 tags", 0)
        basic_tunings.setdefault("Number of H3 tags", 0)
        basic_tunings.setdefault("Number of H4 tags", 0)
        basic_tunings.setdefault("Number of H5 tags", 0)
        basic_tunings.setdefault("Number of H6 tags", 0)
        basic_tunings.setdefault("Number of Images", 0)
        basic_tunings.setdefault("Number of heading tags", 0)
        requirements.setdefault("Title Length", 60)
        requirements.setdefault("Description Length", 160)
        
        # Parse "LSI Keywords" sheet
        lsi_sheet_name = next((s for s in wb.sheetnames if "LSI" in s and "Keywords" in s), None)
        lsi_keywords = {}
        if lsi_sheet_name:
            lsi_sheet = wb[lsi_sheet_name]
            lsi_keywords_data = []
            for row in range(7, lsi_sheet.max_row + 1):
                keyword = lsi_sheet.cell(row=row, column=1).value
                avg = lsi_sheet.cell(row=row, column=2).value
                g_value = lsi_sheet.cell(row=row, column=7).value
                if keyword and avg:
                    try:
                        avg_float = float(avg)
                        g_float = float(g_value) if g_value else 0
                        rounded_g = math.ceil(g_float) if g_float > 0 else 1
                        lsi_keywords_data.append((keyword, rounded_g, g_float))
                    except ValueError:
                        continue
            lsi_keywords_data.sort(key=lambda x: x[2], reverse=True)
            lsi_keywords = {item[0]: item[1] for item in lsi_keywords_data}
        
        # Parse "Entities" sheet
        if "Entities" in wb.sheetnames:
            entities_sheet = wb["Entities"]
            for row in range(4, entities_sheet.max_row + 1):
                entity = entities_sheet.cell(row=row, column=1).value
                if entity:
                    entities.append(str(entity).strip())
        
        # Custom Entities (from user or other logic)
        custom_entities = []
        if "Custom Entities" in wb.sheetnames:
            custom_entities_sheet = wb["Custom Entities"]
            for row in range(2, custom_entities_sheet.max_row + 1):
                entity = custom_entities_sheet.cell(row=row, column=1).value
                if entity:
                    custom_entities.append(str(entity).strip())
        
        # Build typed dataclass
        heading_targets = HeadingTargets.from_dict(basic_tunings)

        seoreq = SEORequirements(
            primary_keyword=primary_keyword,
            variations=variations,
            lsi_keywords=lsi_keywords,
            entities=entities,
            headings=heading_targets,
            word_count=basic_tunings.get("Word Count", 1500),
            images=basic_tunings.get("Number of Images", 0),
            basic_tunings=basic_tunings,
            roadmap_requirements=requirements,
            debug_info=debug_info,
        )

        logger.info("Successfully parsed CORA report for '%s'", primary_keyword)
        return seoreq
    
    except Exception as e:
        logger.error("Error parsing CORA report: %s", str(e))
        raise ParseError(f"Failed to parse CORA report: {str(e)}") from e
